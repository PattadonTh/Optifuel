"""
src/loader.py
=============
Function 1 — Load shipment data and parameters from the Assumption sheet.

Supports both .xlsm and .xlsx files.
Auto-detects demurrage rates if present; falls back to defaults.
"""

import openpyxl
from datetime import datetime


def load_assumption(path: str) -> tuple[list[dict], dict]:
    """
    Read the Assumption sheet from a workbook and return shipments + params.

    Parameters
    ----------
    path : str
        Full path to the .xlsm or .xlsx file.

    Returns
    -------
    shipments : list[dict]
        Sorted by win_start. Each dict has keys:
            sm          (int)      shipment number
            site        (str)      'Gh-1' or 'SPP3'
            win_start   (datetime) window start date
            win_end     (datetime) window end date
            num_window  (int)      window width in days

    params : dict
        Keys:
            n_shipments  (int)   total shipment count
            min_dis      (float) minimum discharge days
            max_dis      (float) maximum discharge days
            allowance    (float) loading allowance days (from NOR+12)
            gh1_rate     (float) Gh-1 demurrage rate USD/day
            spp3_rate    (float) SPP3 demurrage rate USD/day
    """
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb["Assumption"]
    rows = list(ws.iter_rows(values_only=True))

    # ── Default parameters ────────────────────────────────────────────────────
    params = {
        "n_shipments": None,
        "min_dis":     2.0,
        "max_dis":     2.9,
        "allowance":   2.5,
        "gh1_rate":    5500.0,
        "spp3_rate":   2000.0,
    }

    # Maps lowercase label → params key
    _param_map = {
        "number of shipment":  "n_shipments",
        "min discharge":       "min_dis",
        "max discharge":       "max_dis",
        "discharge allowance": "allowance",
        "gh-1 demurrage rate": "gh1_rate",
        "gh1 demurrage rate":  "gh1_rate",
        "spp3 demurrage rate": "spp3_rate",
    }

    # ── Parse parameters ──────────────────────────────────────────────────────
    for row in rows:
        label = str(row[1]).strip().lower() if row[1] else ""
        value = row[3]
        if label in _param_map and isinstance(value, (int, float)):
            params[_param_map[label]] = float(value)

    # ── Parse shipments ───────────────────────────────────────────────────────
    shipments = []
    for row in rows:
        sm_num = row[1]
        # Must be a number (shipment index)
        if sm_num is None or not isinstance(sm_num, (int, float)):
            continue
        # Site must be a string ('Gh-1' or 'SPP3')
        site = row[2]
        if not isinstance(site, str):
            continue
        win_start  = row[3]
        win_end    = row[4]
        num_window = row[5] if row[5] is not None else 10
        # Window start must be a valid date
        if not isinstance(win_start, datetime):
            continue
        shipments.append({
            "sm":         int(sm_num),
            "site":       site.strip(),
            "win_start":  win_start,
            "win_end":    win_end,
            "num_window": int(num_window),
        })

    shipments.sort(key=lambda x: x["win_start"])

    if params["n_shipments"] is None:
        params["n_shipments"] = len(shipments)

    # ── Summary log ───────────────────────────────────────────────────────────
    gh1_count  = sum(1 for s in shipments if s["site"] == "Gh-1")
    spp3_count = sum(1 for s in shipments if s["site"] == "SPP3")
    print(f"[loader] {len(shipments)} shipments loaded  "
          f"(Gh-1: {gh1_count}, SPP3: {spp3_count})")
    print(f"[loader] Params — min_dis={params['min_dis']}d  "
          f"max_dis={params['max_dis']}d  allowance={params['allowance']}d  "
          f"gh1_rate=USD {params['gh1_rate']:,.0f}/d  "
          f"spp3_rate=USD {params['spp3_rate']:,.0f}/d")

    return shipments, params


def load_simulation(path: str) -> dict[int, dict]:
    """
    Read per-ship actual demurrage days from the Simulation sheet.

    Returns
    -------
    dict keyed by sm (int):
        dem_days  (float)  actual demurrage days from the single RAND run
        cost_usd  (float)  dem_days × rate (rate inferred from site)

    Returns empty dict if Simulation sheet does not exist.
    """
    try:
        wb   = openpyxl.load_workbook(path, read_only=True, data_only=True)
        if "Simulation" not in wb.sheetnames:
            return {}
        ws   = wb["Simulation"]
        rows = list(ws.iter_rows(values_only=True))

        # Read rates from summary rows (rows 2-5, col D label / col E value)
        # Fallback to standard defaults
        rates = {"Gh-1": 5500.0, "SPP3": 2000.0}

        # Find header row (contains "Shipment")
        hdr_idx = None
        for i, row in enumerate(rows):
            if row[1] == "Shipment":
                hdr_idx = i
                break
        if hdr_idx is None:
            return {}

        result = {}
        for row in rows[hdr_idx + 1:]:
            sm = row[1]
            if sm is None or not isinstance(sm, (int, float)):
                continue
            site     = str(row[2]).strip() if row[2] else ""
            dem_days = row[12] if isinstance(row[12], (int, float)) else 0.0
            rate     = rates.get(site, 5500.0)
            result[int(sm)] = {
                "dem_days": dem_days,
                "cost_usd": dem_days * rate,
                "site":     site,
            }

        print(f"[loader] Simulation sheet: {len(result)} ships read  "
              f"(total dem days: {sum(v['dem_days'] for v in result.values()):.4f}d  "
              f"cost: USD {sum(v['cost_usd'] for v in result.values()):,.0f})")
        return result

    except Exception as e:
        print(f"[loader] Warning: could not read Simulation sheet — {e}")
        return {}


def load_summary(path: str) -> dict | None:
    """
    Read pre-computed simulation stats from the Summary sheet.

    Returns None if Summary sheet does not exist or has no trial data.

    Returns dict with keys:
        trials          list[tuple]  (trial_num, avg_dis, total_dem, gh1_dem, spp3_dem)
        avg_total_dem   float
        avg_gh1_dem     float
        avg_spp3_dem    float
        sd              float
        conf_95         float
        ci_min          float
        ci_max          float
        n_trials        int
    """
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        if "Summary" not in wb.sheetnames:
            return None
        ws   = wb["Summary"]
        rows = list(ws.iter_rows(values_only=True))

        # Read header stats block (rows 1-4)
        # Layout: row1=labels(Total/Gh-1/SPP3/SD), row2=AVG, row3=MIN, row4=MAX
        # Columns: B=label, C=total, D=gh1, E=spp3, H=stat_label, I=stat_val
        avg_total = rows[1][2]
        avg_gh1   = rows[1][3]
        avg_spp3  = rows[1][4]
        sd        = rows[0][8]   # SD value in row 1, col I
        conf_95   = rows[1][8]   # 95% confident value
        ci_min    = rows[2][8]   # 95% min
        ci_max    = rows[3][8]   # 95% max

        # Read trial rows (row 8 onward): B=trial, C=avg_dis, D=total, E=gh1, F=spp3
        trials = []
        for row in rows[7:]:
            t = row[1]
            if t is None or not isinstance(t, (int, float)):
                continue
            trials.append((int(t), row[2], row[3], row[4], row[5]))

        if not trials:
            return None

        result = {
            "trials":        trials,
            "avg_total_dem": avg_total,
            "avg_gh1_dem":   avg_gh1,
            "avg_spp3_dem":  avg_spp3,
            "sd":            sd,
            "conf_95":       conf_95,
            "ci_min":        ci_min,
            "ci_max":        ci_max,
            "n_trials":      len(trials),
        }

        print(f"[loader] Summary sheet: {len(trials)} trials  "
              f"AVG dem={avg_total:.4f}d  "
              f"(Gh-1={avg_gh1:.4f}d, SPP3={avg_spp3:.4f}d)  "
              f"SD={sd:.4f}")
        return result

    except Exception as e:
        print(f"[loader] Warning: could not read Summary sheet — {e}")
        return None
