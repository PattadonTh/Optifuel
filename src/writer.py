"""
src/writer.py
=============
Function 3 — Write optimization results to Excel.

Output workbook has 3 sheets:
  1. Summary            — Monte Carlo trial table + stats block
                          (matches Estimate_Demurrage.xlsm layout exactly)
  2. Optimization Summary — Model parameters, cost results, gap analysis
  3. Window Schedule    — All ships: baseline vs optimised dates + per-ship cost
"""

import numpy as np
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta


# ── Style helpers (module-level so all _build_* functions share them) ─────────

def _font(bold=False, size=10, color="FF000000", italic=False):
    return Font(name="Arial", bold=bold, size=size, color=color, italic=italic)

def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def _aln(h="center", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

_thin = Side(style="thin",   color="FFB8CCE4")
_med  = Side(style="medium", color="FF1F497D")

def _bdr():  return Border(top=_thin, bottom=_thin, left=_thin, right=_thin)
def _hbdr(): return Border(top=_med,  bottom=_med,  left=_med,  right=_med)

# Pre-built font / fill constants
F_HDR   = _font(bold=True,  color="FFFFFFFF")
F_BODY  = _font()
F_BOLD  = _font(bold=True)
F_TITLE = _font(bold=True,  size=12, color="FF1F497D")
F_GRAY  = _font(size=9,     italic=True, color="FF808080")
F_RED   = _font(bold=True,  color="FF9C0006")
F_GREEN = _font(bold=True,  color="FF375623")

FILL_HDR   = _fill("FF1F497D")
FILL_SPP3  = _fill("FFE2EFDA")
FILL_SPP3A = _fill("FFF0F7EC")
FILL_GH1   = _fill("FFDCE6F1")
FILL_GH1A  = _fill("FFEFF5FB")
FILL_SUM   = _fill("FFEDEDED")
FILL_NOTE  = _fill("FFFFF2CC")
FILL_OPT   = _fill("FFEBF5FF")

DATE_FMT = "DD-MMM-YYYY"
USD_FMT  = "#,##0"
N4_FMT   = "#,##0.0000"
N2_FMT   = "#,##0.00"
PCT_FMT  = "0.0%"

COL_MAP = {c: i for i, c in enumerate("ABCDEFGHIJKLMNOP", 1)}


def _hdr(ws, row, col, val, width=None):
    """Write a header cell (dark blue bg, white bold text)."""
    c = ws.cell(row=row, column=col, value=val)
    c.font = F_HDR; c.fill = FILL_HDR
    c.alignment = _aln(); c.border = _hbdr()
    if width:
        ws.column_dimensions[get_column_letter(col)].width = width
    return c


def _cell(ws, row, col, val, fmt=None, fnt=None, fl=None, al=None):
    """Write a data cell with optional format / font / fill / alignment."""
    c = ws.cell(row=row, column=col, value=val)
    c.font      = fnt if fnt else F_BODY
    c.fill      = fl  if fl  else _fill("00000000")   # transparent
    c.alignment = al  if al  else _aln()
    c.border    = _bdr()
    if fmt:
        c.number_format = fmt
    return c


# ── Sheet builders ────────────────────────────────────────────────────────────

def _build_summary(
    wb,
    gh1_od, gh1_oc,
    spp3_od, spp3_oc,
    params,
    n_eval,
    actual_sim,
):
    """
    Sheet 1 — Summary.
    Baseline  = ค่าจริงจาก Simulation sheet (single run ที่ run มาแล้ว)
    Optimised = E[cost] จาก Monte Carlo หลัง Lagrange optimizer
    แยกตาม site (Gh-1 / SPP3 / Total)
    """
    ws = wb.active
    ws.title = "Summary"
    ws.sheet_view.showGridLines = False

    # Baseline — ใช้ค่าจริงจาก Simulation sheet โดยตรง
    act_gh1_days  = sum(v["dem_days"] for v in actual_sim.values() if v["site"] == "Gh-1")
    act_spp3_days = sum(v["dem_days"] for v in actual_sim.values() if v["site"] == "SPP3")
    act_gh1_cost  = sum(v["cost_usd"] for v in actual_sim.values() if v["site"] == "Gh-1")
    act_spp3_cost = sum(v["cost_usd"] for v in actual_sim.values() if v["site"] == "SPP3")
    baseline_label = "ค่าจริงจาก Simulation sheet (single run)"

    # Aggregate optimised by site
    opt_gh1_days  = sum(gh1_od)
    opt_spp3_days = sum(spp3_od)
    opt_gh1_cost  = sum(gh1_oc)
    opt_spp3_cost = sum(spp3_oc)

    # Title
    ws.merge_cells("B2:I2")
    ws["B2"] = "Demurrage Summary — Actual vs Optimised"
    ws["B2"].font = _font(bold=True, size=12, color="FF1F497D")
    ws["B2"].alignment = _aln("left")

    ws.merge_cells("B3:I3")
    ws["B3"] = (f"Gh-1: USD {params['gh1_rate']:,.0f}/day  ·  "
                f"SPP3: USD {params['spp3_rate']:,.0f}/day  ·  "
                f"Baseline = ค่าจริงจาก Simulation sheet  ·  "
                f"Optimised = E[cost] จาก Monte Carlo {n_eval:,} sims")
    ws["B3"].font = _font(size=9, italic=True, color="FF808080")

    # Column headers row 5
    col_hdrs = ["Site",
                "Baseline dem days", "Baseline cost USD",
                "Optimised dem days", "Optimised cost USD",
                "Saving (days)", "Saving (USD)", "Saving (%)"]
    col_widths = [10, 22, 22, 22, 22, 16, 16, 12]
    for ci, (h, w) in enumerate(zip(col_hdrs, col_widths), 2):
        c = ws.cell(row=5, column=ci, value=h)
        c.font = F_HDR; c.fill = FILL_HDR
        c.alignment = _aln(wrap=True); c.border = _hbdr()
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[5].height = 36

    # Data rows
    rows_data = [
        ("Gh-1",  act_gh1_days,  act_gh1_cost,  opt_gh1_days,  opt_gh1_cost,  FILL_GH1),
        ("SPP3",  act_spp3_days, act_spp3_cost, opt_spp3_days, opt_spp3_cost, FILL_SPP3),
        ("Total", act_gh1_days + act_spp3_days,
                  act_gh1_cost  + act_spp3_cost,
                  opt_gh1_days  + opt_spp3_days,
                  opt_gh1_cost  + opt_spp3_cost, FILL_SUM),
    ]
    for ri, (site, ad, ac, od, oc, fl) in enumerate(rows_data, 6):
        saving_days = ad - od
        saving_cost = ac - oc
        saving_pct  = saving_cost / ac if ac else 0.0
        is_total    = (site == "Total")
        vals = [site, ad, ac, od, oc, saving_days, saving_cost, saving_pct]
        fmts = [None, N4_FMT, USD_FMT, N4_FMT, USD_FMT, N4_FMT, USD_FMT, "0.0%"]
        for ci, (val, fmt) in enumerate(zip(vals, fmts), 2):
            fnt = F_BOLD if is_total else (
                  F_GREEN if ci in (8, 9) and isinstance(val, (int, float)) and val > 0
                  else F_BODY)
            _cell(ws, ri, ci, val, fmt=fmt, fnt=fnt, fl=fl)
        ws.row_dimensions[ri].height = 20

    # Note



def _build_window_schedule(
    wb, shipments, opt_result,
    gh1_bd, gh1_bc, gh1_od, gh1_oc,
    spp3_bd, spp3_bc, spp3_od, spp3_oc,
    BASE_DATE,
    actual_sim: dict = None,
):
    actual_sim = actual_sim or {}
    """
    Sheet 3 — Window Schedule.
    All ships sorted chronologically: baseline vs optimised dates, shift,
    gap to next ship, per-ship E[dem days], E[cost], and cost saving.
    """
    ws = wb.create_sheet("Window Schedule")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("B2:O2")
    ws["B2"] = "Window Schedule — Baseline vs Optimised"
    ws["B2"].font = F_TITLE; ws["B2"].alignment = _aln("left")

    # ── Headers row 4 ─────────────────────────────────────────────────────────
    hdrs   = ["SM#", "Site",
              "Baseline\nwin start", "Baseline\nwin end",
              "Optimised\nwin start", "Optimised\nwin end",
              "Window shift\n(days)",
              "Actual dem days", "Actual cost USD",
              "Optimised dem days", "Optimised cost USD",
              "Cost saving\n(USD)"]
    widths = [8, 8, 16, 16, 16, 16, 13, 22, 22, 22, 22, 16]
    for ci, (h, w) in enumerate(zip(hdrs, widths), 2):
        c = ws.cell(row=4, column=ci, value=h)
        c.font = F_HDR; c.fill = FILL_HDR
        c.alignment = _aln(wrap=True); c.border = _hbdr()
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[4].height = 36

    # ── Lookup maps ───────────────────────────────────────────────────────────
    gh1_opt_map  = {s["sm"]: opt_result["gh1_opt"][i]
                    for i, s in enumerate(opt_result["gh1_ships"])}
    spp3_opt_map = {s["sm"]: opt_result["spp3_opt"][i]
                    for i, s in enumerate(opt_result["spp3_ships"])}

    def _gap_maps(ships, base_arr, opt_arr):
        bg = {ships[i]["sm"]: base_arr[i + 1] - base_arr[i]
              for i in range(len(ships) - 1)}
        og = {ships[i]["sm"]: opt_arr[i + 1] - opt_arr[i]
              for i in range(len(ships) - 1)}
        bg[ships[-1]["sm"]] = None
        og[ships[-1]["sm"]] = None
        return bg, og

    gh1_bgm,  gh1_ogm  = _gap_maps(opt_result["gh1_ships"],
                                    opt_result["gh1_base"], opt_result["gh1_opt"])
    spp3_bgm, spp3_ogm = _gap_maps(opt_result["spp3_ships"],
                                    opt_result["spp3_base"], opt_result["spp3_opt"])

    gh1_per  = {s["sm"]: (gh1_bd[i],  gh1_bc[i],  gh1_od[i],  gh1_oc[i])
                for i, s in enumerate(opt_result["gh1_ships"])}
    spp3_per = {s["sm"]: (spp3_bd[i], spp3_bc[i], spp3_od[i], spp3_oc[i])
                for i, s in enumerate(opt_result["spp3_ships"])}

    # ── Data rows ─────────────────────────────────────────────────────────────
    spp3_i = gh1_i = 0
    for ri, sm in enumerate(shipments, 5):
        is_spp3 = sm["site"] == "SPP3"
        if is_spp3:
            fl     = FILL_SPP3 if spp3_i % 2 == 0 else FILL_SPP3A; spp3_i += 1
            opt_d  = spp3_opt_map[sm["sm"]]
            bg_    = spp3_bgm[sm["sm"]]; og_ = spp3_ogm[sm["sm"]]
            bd, bc, od, oc = spp3_per[sm["sm"]]
        else:
            fl     = FILL_GH1 if gh1_i % 2 == 0 else FILL_GH1A; gh1_i += 1
            opt_d  = gh1_opt_map[sm["sm"]]
            bg_    = gh1_bgm[sm["sm"]]; og_ = gh1_ogm[sm["sm"]]
            bd, bc, od, oc = gh1_per[sm["sm"]]

        base_ws = sm["win_start"]
        opt_ws  = BASE_DATE + timedelta(days=float(opt_d))
        shift    = float(opt_d) - (base_ws - BASE_DATE).days

        # Shift cell colour: yellow = moved earlier, blue = moved later, grey = unchanged
        shift_fl = (FILL_NOTE if shift < -1 else
                    FILL_OPT  if shift >  1 else
                    FILL_SUM)

        # Actual from Simulation sheet
        act      = actual_sim.get(sm["sm"], {})
        act_dem  = act.get("dem_days", None)
        act_cost = act.get("cost_usd", None)

        # Cost saving = Actual cost (simulation) - Optimised cost (MC model)
        saving   = act_cost - oc if act_cost is not None else None

        vals = [
            sm["sm"], sm["site"],
            base_ws,  base_ws + timedelta(days=9),
            opt_ws,   opt_ws  + timedelta(days=9),
            round(shift, 1),
            round(act_dem,  4) if act_dem  is not None else "—",
            round(act_cost, 0) if act_cost is not None else "—",
            round(od, 4), round(oc),
            round(saving)      if saving   is not None else "—",
        ]
        fmts = [None, None,
                DATE_FMT, DATE_FMT, DATE_FMT, DATE_FMT,
                "+0.0;-0.0;0.0",
                N4_FMT, USD_FMT,
                N4_FMT, USD_FMT, USD_FMT]

        # col 2=SM#  3=Site  4=base_ws  5=base_we  6=opt_ws  7=opt_we
        #     8=shift  9=act_dem  10=act_cost  11=opt_dem  12=opt_cost  13=saving
        for ci, (val, fmt) in enumerate(zip(vals, fmts), 2):
            cell_fl = shift_fl if ci == 8 else fl
            fnt = (F_GREEN if ci == 13 and isinstance(val, (int, float)) and val > 0
                   else F_RED if ci == 13 and isinstance(val, (int, float)) and val < -1
                   else F_BODY)
            _cell(ws, ri, ci, val, fmt=fmt, fnt=fnt, fl=cell_fl)
        ws.row_dimensions[ri].height = 17

    ws.freeze_panes = "B5"


# ── Public entry point ────────────────────────────────────────────────────────

def write_excel(
    output_path: str,
    shipments:   list[dict],
    params:      dict,
    opt_result:  dict,
    gh1_bd:  list[float], gh1_bc:  list[float],
    gh1_od:  list[float], gh1_oc:  list[float],
    spp3_bd: list[float], spp3_bc: list[float],
    spp3_od: list[float], spp3_oc: list[float],
    BASE_DATE: datetime,
    n_eval:        int  = 2000,
    actual_sim:    dict = None,
    gh1_base_tot   = None,
    spp3_base_tot  = None,
    total_base_tot = None,
    gh1_opt_tot    = None,
    spp3_opt_tot   = None,
    total_opt_tot  = None,
):
    """
    Write the 3-sheet result workbook to output_path.

    Called by main.py after load + optimize + per-ship simulation are done.
    """
    actual_sim = actual_sim or {}
    wb = openpyxl.Workbook()

    _build_summary(
        wb,
        gh1_od, gh1_oc,
        spp3_od, spp3_oc,
        params,
        n_eval,
        actual_sim=actual_sim or {},
    )
    _build_window_schedule(
        wb, shipments, opt_result,
        gh1_bd, gh1_bc, gh1_od, gh1_oc,
        spp3_bd, spp3_bc, spp3_od, spp3_oc,
        BASE_DATE,
        actual_sim=actual_sim or {},
    )

    _build_mc_sheet(
        wb,
        n_eval         = n_eval,
        gh1_base_tot   = gh1_base_tot,
        spp3_base_tot  = spp3_base_tot,
        total_base_tot = total_base_tot,
        gh1_opt_tot    = gh1_opt_tot,
        spp3_opt_tot   = spp3_opt_tot,
        total_opt_tot  = total_opt_tot,
    )

    wb.save(output_path)
    print(f"[writer] Saved → {output_path}")


def _build_mc_sheet(
    wb,
    n_eval,
    gh1_base_tot, spp3_base_tot, total_base_tot,
    gh1_opt_tot,  spp3_opt_tot,  total_opt_tot,
    **kwargs,
):
    """
    Sheet 3 — MC Simulation
    Raw trial totals: 2,000 runs × (Baseline + Optimised) per site + grand total.
    No background colour.
    """
    import numpy as np

    ws = wb.create_sheet("MC Simulation")
    ws.sheet_view.showGridLines = False

    USD_FMT = "#,##0"
    INT_FMT = "#,##0"

    # ── Title ─────────────────────────────────────────────────────────────
    ws.merge_cells("B2:I2")
    ws["B2"] = "MC Simulation — Raw Trial Totals"
    ws["B2"].font = _font(bold=True, size=12, color="FF1F497D")
    ws["B2"].alignment = _aln("left")

    ws.merge_cells("B3:I3")
    ws["B3"] = f"{n_eval:,} simulation runs  ·  Baseline = original window starts  ·  Optimised = Lagrange optimizer"
    ws["B3"].font = _font(size=9, italic=True, color="FF808080")

    # ── Column widths ─────────────────────────────────────────────────────
    for col, w in [("B",10),("C",18),("D",18),("E",18),("F",18),("G",18),("H",18)]:
        ws.column_dimensions[col].width = w

    # ── Headers ───────────────────────────────────────────────────────────
    r = 5
    hdrs = ["Trial #",
            "Baseline Gh-1", "Baseline SPP3", "Baseline Total",
            "Optimised Gh-1","Optimised SPP3","Optimised Total"]
    for ci, h in enumerate(hdrs, 2):
        _hdr(ws, r, ci, h)
    ws.row_dimensions[r].height = 20
    r += 1

    # ── Data rows — no fill ───────────────────────────────────────────────
    for i in range(n_eval):
        vals = [i + 1,
                gh1_base_tot[i], spp3_base_tot[i], total_base_tot[i],
                gh1_opt_tot[i],  spp3_opt_tot[i],  total_opt_tot[i]]
        fmts = [INT_FMT] + [USD_FMT] * 6
        for ci, (v, fmt) in enumerate(zip(vals, fmts), 2):
            c = ws.cell(row=r, column=ci, value=v)
            c.number_format = fmt
            c.font          = F_BODY
            c.alignment     = _aln()
            c.border        = _bdr()
        ws.row_dimensions[r].height = 15
        r += 1

    ws.freeze_panes = "B6"
    print(f"[writer] MC Simulation sheet: {n_eval:,} rows written")

